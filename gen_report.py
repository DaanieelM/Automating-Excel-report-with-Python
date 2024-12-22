import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

# Combine files to Excel, all Excel files in the folder will be combine
path = r'C:\Users\Daniel\Desktop\Python\Project\data\\'
files = os.listdir(path)
combined = pd.DataFrame()
for file in files:
    df = pd.read_excel(path+file)
    combined = pd.concat([combined, df], ignore_index=True)
    
# combined.to_excel(r'C:\Users\Daniel\Desktop\Combined.xlsx', index=False)

# Grouping data
summary = pd.pivot_table(
    data = combined,
    index = 'Salesperson',
    values = ['Amount', 'Costs', 'Profit'],
    aggfunc = 'sum'
    
)

# Export grouped data to Excel
summary.to_excel(r'C:\Users\Daniel\Desktop\Summary.xlsx')

# Insert and customize data
wb = load_workbook(r'C:\Users\Daniel\Desktop\Summary.xlsx')
ws = wb['Sheet1']

print(ws['A1'].value)

ws.insert_rows(0, 3)
ws['A1'].value = 'Sales by Person'
ws['A2'].value = 'Automated report'

ws['A1'].style = 'Title'
ws['A2'].style = 'Headline 2'

for cell in range(5, ws.max_row+1):
    ws[f'B{cell}'].style= "Currency"

for cell in range(5, ws.max_row+1):
    ws[f'C{cell}'].style= "Currency"

for cell in range(5, ws.max_row+1):
    ws[f'D{cell}'].style= "Currency"
    
# Add Charts to Worbook
data = Reference(ws, min_col=2, max_col=2, min_row=5, max_row=ws.max_row)
categories = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=ws.max_row)

chart = BarChart()
chart.add_data(data)
chart.y_axis.title = 'Amount'
chart.x_axis.title = 'Persons'
chart.set_categories(categories)
chart.title = 'Sales by Person'
ws.add_chart(chart, anchor='F4')
chart.legend = None

data2 = Reference(ws, min_col=3, max_col=3, min_row=5, max_row=ws.max_row)
categories = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=ws.max_row)

chart = BarChart()
chart.add_data(data2)
chart.y_axis.title = 'Amount'
chart.x_axis.title = 'Persons'
chart.set_categories(categories)
chart.title = 'Costs by Person'
ws.add_chart(chart, anchor='F19')
chart.legend = None

data3 = Reference(ws, min_col=4, max_col=4, min_row=5, max_row=ws.max_row)
categories = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=ws.max_row)

chart = BarChart()
chart.add_data(data3)
chart.y_axis.title = 'Amount'
chart.x_axis.title = 'Persons'
chart.set_categories(categories)
chart.title = 'Profit by Persons'
ws.add_chart(chart, anchor='P19')
chart.legend = None

from copy import deepcopy
chart2 = deepcopy(chart)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horiznotal Bar Chart"
ws.add_chart(chart2, "P4")
chart.legend = None

# Save report to Excel file
wb.save(r'C:\Users\Daniel\Desktop\Summary.xlsx')
